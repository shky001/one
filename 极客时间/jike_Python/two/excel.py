# Excel 格式
# pip install openpyxl

from openpyxl import Workbook
import datetime

wb = Workbook()     # wb为整个Excel文件
ws = wb.active      # ws为sheet表

ws['A1'] = 42
ws.append([1,2,3])

ws['A2'] = datetime.datetime.now()

wb.save("sample.xlsx")




